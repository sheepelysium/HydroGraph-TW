# -*- coding: utf-8 -*-
"""提取集水區資料並與河川建立關聯"""
import pandas as pd
from dbfread import DBF
from pathlib import Path

def read_watershed_data():
    """讀取集水區 DBF 資料"""

    print("讀取集水區資料...")
    dbf_path = 'file/110年度全臺839子集水區範圍圖_UTF8.dbf'
    table = DBF(dbf_path, encoding='utf-8')

    # 轉換為 DataFrame
    df = pd.DataFrame(iter(table))

    print(f"集水區數量: {len(df)}")
    print(f"流域數量: {df['BASIN_ID'].nunique()}")

    return df

def match_watersheds_to_rivers(watersheds_df, rivers_df):
    """將集水區與河川建立關聯"""

    print("\n建立集水區與河川的關聯...")

    # 建立河川代碼到河川名稱的映射（使用前4碼）
    river_mapping = {}
    for idx, row in rivers_df.iterrows():
        river_code = str(row['河川代碼'])
        river_name = row['河川名稱']
        main_stream = row['主流水系']
        level = row['階層']

        # 提取前4碼作為流域代碼
        if len(river_code) >= 4:
            basin_code = river_code[:4]
            if basin_code not in river_mapping:
                river_mapping[basin_code] = []
            river_mapping[basin_code].append({
                '河川名稱': river_name,
                '河川代碼': river_code,
                '主流水系': main_stream,
                '階層': level
            })

    print(f"河川流域代碼（前4碼）數量: {len(river_mapping)}")

    # 為每個集水區匹配河川
    watersheds_df['流域代碼'] = watersheds_df['WS_ID'].apply(lambda x: str(x)[:4] if pd.notna(x) else None)

    # 統計匹配情況
    matched_count = 0
    watersheds_df['關聯河川數量'] = 0
    watersheds_df['主要河川'] = None

    for idx, row in watersheds_df.iterrows():
        basin_code = row['流域代碼']
        if basin_code in river_mapping:
            rivers = river_mapping[basin_code]
            matched_count += 1
            watersheds_df.at[idx, '關聯河川數量'] = len(rivers)
            # 選擇階層最小的（主流）作為主要河川
            main_river = min(rivers, key=lambda x: x['階層'])
            watersheds_df.at[idx, '主要河川'] = main_river['河川名稱']

    print(f"能匹配到河川的集水區: {matched_count} / {len(watersheds_df)} ({matched_count/len(watersheds_df)*100:.1f}%)")

    # 統計每個流域的集水區數量
    basin_stats = watersheds_df.groupby('BASIN_NAME').agg({
        'WS_ID': 'count',
        'AREA_M2': 'sum',
        '關聯河川數量': 'sum'
    }).rename(columns={
        'WS_ID': '集水區數量',
        'AREA_M2': '總面積(m2)',
        '關聯河川數量': '關聯河川總數'
    }).sort_values('集水區數量', ascending=False)

    return watersheds_df, basin_stats, river_mapping

def generate_watershed_report(watersheds_df, basin_stats, river_mapping, rivers_df):
    """產生集水區分析報表"""

    output_path = Path('data/集水區分析報表.xlsx')
    print(f"\n正在產生報表: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 工作表1: 摘要統計
        summary_data = []

        summary_data.append({'項目': '【集水區基本統計】', '數量': '', '備註': ''})
        summary_data.append({
            '項目': '總集水區數',
            '數量': len(watersheds_df),
            '備註': '839個子集水區'
        })
        summary_data.append({
            '項目': '總流域數',
            '數量': watersheds_df['BASIN_ID'].nunique(),
            '備註': '主要流域數量'
        })
        summary_data.append({
            '項目': '總面積',
            '數量': f"{watersheds_df['AREA_M2'].sum()/1e6:.2f} km²",
            '備註': '全台灣集水區總面積'
        })

        summary_data.append({'項目': '', '數量': '', '備註': ''})
        summary_data.append({'項目': '【與河川關聯統計】', '數量': '', '備註': ''})

        matched_watersheds = watersheds_df[watersheds_df['關聯河川數量'] > 0]
        summary_data.append({
            '項目': '已關聯河川的集水區',
            '數量': len(matched_watersheds),
            '備註': f'{len(matched_watersheds)/len(watersheds_df)*100:.1f}%'
        })
        summary_data.append({
            '項目': '未關聯河川的集水區',
            '數量': len(watersheds_df) - len(matched_watersheds),
            '備註': f'{(len(watersheds_df) - len(matched_watersheds))/len(watersheds_df)*100:.1f}%'
        })
        summary_data.append({
            '項目': '流域代碼匹配數',
            '數量': len(river_mapping),
            '備註': '可與河川代碼前4碼匹配'
        })

        summary_data.append({'項目': '', '數量': '', '備註': ''})
        summary_data.append({'項目': '【前10大流域（按集水區數量）】', '數量': '', '備註': ''})

        for basin_name, stats in basin_stats.head(10).iterrows():
            summary_data.append({
                '項目': f'  {basin_name}',
                '數量': f"{stats['集水區數量']} 個",
                '備註': f"{stats['總面積(m2)']/1e6:.1f} km²"
            })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='摘要統計', index=False)

        # 工作表2: 集水區列表
        ws_cols = ['WS_ID', 'WS_NAME', 'BASIN_ID', 'BASIN_NAME', 'AREA_M2',
                   '流域代碼', '關聯河川數量', '主要河川', 'BRANCH']
        ws_output = watersheds_df[ws_cols].copy()
        ws_output = ws_output.sort_values(['BASIN_ID', 'WS_ID'])
        ws_output.to_excel(writer, sheet_name='集水區列表', index=False)

        # 工作表3: 流域統計
        basin_stats_output = basin_stats.copy()
        basin_stats_output['平均集水區面積(km2)'] = basin_stats_output['總面積(m2)'] / basin_stats_output['集水區數量'] / 1e6
        basin_stats_output['總面積(km2)'] = basin_stats_output['總面積(m2)'] / 1e6
        basin_stats_output = basin_stats_output.drop('總面積(m2)', axis=1)
        basin_stats_output.to_excel(writer, sheet_name='流域統計', index=True)

        # 工作表4: 集水區-河川關聯表
        relation_data = []
        for idx, row in watersheds_df[watersheds_df['關聯河川數量'] > 0].iterrows():
            basin_code = row['流域代碼']
            if basin_code in river_mapping:
                for river in river_mapping[basin_code]:
                    relation_data.append({
                        '集水區ID': row['WS_ID'],
                        '集水區名稱': row['WS_NAME'],
                        '流域代碼': basin_code,
                        '流域名稱': row['BASIN_NAME'],
                        '河川代碼': river['河川代碼'],
                        '河川名稱': river['河川名稱'],
                        '河川階層': river['階層'],
                        '主流水系': river['主流水系']
                    })

        if relation_data:
            relation_df = pd.DataFrame(relation_data)
            relation_df = relation_df.sort_values(['集水區ID', '河川階層'])
            relation_df.to_excel(writer, sheet_name='集水區-河川關聯', index=False)

    # 設定凍結窗格
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
    wb.save(output_path)

    print(f"\n報表已產生: {output_path}")
    print("\n報表內容:")
    print("  - 工作表1: 摘要統計")
    print(f"  - 工作表2: 集水區列表 ({len(watersheds_df)} 個)")
    print(f"  - 工作表3: 流域統計 ({len(basin_stats)} 個)")
    print(f"  - 工作表4: 集水區-河川關聯 ({len(relation_data)} 筆關係)")

    # 顯示摘要
    print("\n" + "=" * 80)
    print("摘要統計")
    print("=" * 80)
    try:
        print(summary_df.to_string(index=False))
    except UnicodeEncodeError:
        print("（摘要包含特殊字元，請查看報表檔案）")

def main():
    print("=" * 80)
    print("集水區資料提取與分析")
    print("=" * 80)

    # 1. 讀取集水區資料
    watersheds = read_watershed_data()

    # 2. 讀取河川資料
    print("\n讀取河川資料...")
    rivers = pd.read_excel('data/河川關係_完整版.xlsx')
    print(f"河川數量: {len(rivers)}")

    # 3. 建立關聯
    watersheds, basin_stats, river_mapping = match_watersheds_to_rivers(watersheds, rivers)

    # 4. 產生報表
    generate_watershed_report(watersheds, basin_stats, river_mapping, rivers)

    print("\n" + "=" * 80)
    print("完成！")
    print("=" * 80)

if __name__ == "__main__":
    main()
