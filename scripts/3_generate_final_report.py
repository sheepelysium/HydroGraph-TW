# -*- coding: utf-8 -*-
"""產生完整的配對報表（使用改進的匹配邏輯）"""
import pandas as pd
import re
from pathlib import Path

def extract_river_names(text):
    """從河川名稱中提取所有可能的名稱（主名稱+括號內的別名）

    例如：
    - '乾溪(里仁溪)' -> ['乾溪', '里仁溪']
    - '那都魯薩溪（老人溪）' -> ['那都魯薩溪', '老人溪']
    - '東興坑溪【東坑溪】' -> ['東興坑溪', '東坑溪']
    """
    if pd.isna(text) or text == '' or text == 'nan':
        return []

    text = str(text).strip()
    names = []

    # 提取主名稱（移除所有括號及其內容）
    main_name = re.sub(r'[\[\]【】\(\)（）].*?[\]\】\)\）]', '', text).strip()
    if main_name:
        names.append(main_name)

    # 提取所有括號內的內容
    brackets = re.findall(r'[\[\【\(\（](.*?)[\]\】\)\）]', text)
    for bracket_content in brackets:
        bracket_content = bracket_content.strip()
        if bracket_content:
            names.append(bracket_content)

    return names

def build_river_name_mapping(rivers_df):
    """建立河川名稱到河川資訊的映射（包含所有別名）

    Returns:
        dict: {河川名稱變體: (原始河川名稱, 河川代碼, DataFrame索引)}
    """
    mapping = {}

    for idx, row in rivers_df.iterrows():
        original_name = row['河川名稱']
        river_code = row['河川代碼']

        # 提取所有可能的名稱
        all_names = extract_river_names(original_name)

        for name in all_names:
            if name not in mapping:
                mapping[name] = (original_name, river_code, idx)

    return mapping

def match_station_to_river(station_river, river_mapping):
    """將測站的河川名稱匹配到河川資料

    Args:
        station_river: 測站的河川名稱
        river_mapping: 河川名稱映射字典

    Returns:
        tuple: (是否匹配, 匹配的河川名稱, 河川代碼, 匹配方式)
    """
    if pd.isna(station_river) or station_river == '' or station_river == 'nan':
        return (False, None, None, '河川欄位為空')

    station_river = str(station_river).strip()

    # 1. 直接匹配
    if station_river in river_mapping:
        matched_name, river_code, _ = river_mapping[station_river]
        return (True, matched_name, river_code, '直接匹配')

    # 2. 提取測站河川的所有可能名稱並嘗試匹配
    station_names = extract_river_names(station_river)
    for name in station_names:
        if name in river_mapping:
            matched_name, river_code, _ = river_mapping[name]
            return (True, matched_name, river_code, '別名匹配')

    # 3. 無法匹配
    return (False, None, None, '無法匹配')

def generate_complete_report_improved():
    """產生完整配對報表（使用改進的匹配邏輯）"""

    print("讀取資料...")
    stations = pd.read_excel('data/測站資料_水位與氣象.xlsx')
    rivers = pd.read_excel('data/河川關係_完整版.xlsx')

    print(f"測站數量: {len(stations)}")
    print(f"河川數量: {len(rivers)}")

    # 建立河川名稱映射
    print("\n建立河川名稱映射（包含別名）...")
    river_mapping = build_river_name_mapping(rivers)
    print(f"河川名稱變體總數: {len(river_mapping)}")

    # 對測站進行匹配
    print("執行改進的匹配...")
    match_results = []
    for idx, row in stations.iterrows():
        station_river = row.get('河川', None)
        is_matched, matched_name, river_code, match_type = match_station_to_river(
            station_river, river_mapping
        )
        match_results.append({
            'is_matched': is_matched,
            'matched_river': matched_name,
            'river_code': river_code,
            'match_type': match_type
        })

    # 添加到資料框
    stations['能否配對'] = [r['is_matched'] for r in match_results]
    stations['匹配的河川'] = [r['matched_river'] for r in match_results]
    stations['河川代碼'] = [r['river_code'] for r in match_results]
    stations['匹配方式'] = [r['match_type'] for r in match_results]

    matched_stations = stations[stations['能否配對']].copy()
    unmatched_stations = stations[~stations['能否配對']].copy()

    # 建立河川與測站的對應關係（使用匹配的河川名稱）
    rivers['測站數量'] = rivers['河川名稱'].apply(
        lambda x: len(matched_stations[matched_stations['匹配的河川'] == x])
    )
    rivers['有測站'] = rivers['測站數量'] > 0

    rivers_without_stations_df = rivers[rivers['測站數量'] == 0].copy()
    rivers_with_stations_df = rivers[rivers['測站數量'] > 0].copy()

    print(f"\n能配對的測站: {len(matched_stations)} ({len(matched_stations)/len(stations)*100:.1f}%)")
    print(f"無法配對的測站: {len(unmatched_stations)} ({len(unmatched_stations)/len(stations)*100:.1f}%)")
    print(f"\n已配對的河川: {len(rivers_with_stations_df)} / {len(rivers)} ({len(rivers_with_stations_df)/len(rivers)*100:.1f}%)")
    print(f"未配對的河川: {len(rivers_without_stations_df)} / {len(rivers)} ({len(rivers_without_stations_df)/len(rivers)*100:.1f}%) [有待確認]")

    # 匹配方式統計
    print(f"\n匹配方式統計:")
    for match_type, count in matched_stations['匹配方式'].value_counts().items():
        print(f"  {match_type}: {count} ({count/len(matched_stations)*100:.1f}%)")

    # 分類無法配對的測站
    def classify_reason(row):
        river = row['河川']
        if pd.isna(river) or river == '' or river == 'nan':
            return '河川欄位為空'
        elif '排水' in river:
            return '排水系統'
        elif '圳' in river or '溝' in river:
            return '灌溉渠道'
        elif river in ['0000', '??']:
            return '資料錯誤'
        else:
            return '其他（可能是小支流）'

    unmatched_stations['無法配對原因'] = unmatched_stations.apply(classify_reason, axis=1)

    # 按主流水系和階層排序
    rivers_without_stations_df = rivers_without_stations_df.sort_values(['主流水系', '階層', '河川名稱'])
    rivers_with_stations_df = rivers_with_stations_df.sort_values(['測站數量', '主流水系'], ascending=[False, True])

    # 統計報表
    summary_data = []

    # 1. 測站配對統計
    summary_data.append({'項目': '【測站配對統計】', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({
        '項目': '總測站數',
        '數量': len(stations),
        '百分比': '100.0%',
        '備註': ''
    })
    summary_data.append({
        '項目': '能配對的測站',
        '數量': len(matched_stations),
        '百分比': f'{len(matched_stations)/len(stations)*100:.1f}%',
        '備註': '可連結到河川資料'
    })
    summary_data.append({
        '項目': '無法配對的測站',
        '數量': len(unmatched_stations),
        '百分比': f'{len(unmatched_stations)/len(stations)*100:.1f}%',
        '備註': '見「無法配對的測站」工作表'
    })

    # 1.5. 匹配方式統計
    summary_data.append({'項目': '', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({'項目': '【匹配方式統計】', '數量': '', '百分比': '', '備註': ''})
    for match_type, count in matched_stations['匹配方式'].value_counts().items():
        summary_data.append({
            '項目': f'  {match_type}',
            '數量': count,
            '百分比': f'{count/len(matched_stations)*100:.1f}%',
            '備註': ''
        })

    # 2. 河川配對統計
    summary_data.append({'項目': '', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({'項目': '【河川配對統計】', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({
        '項目': '總河川數',
        '數量': len(rivers),
        '百分比': '100.0%',
        '備註': ''
    })
    summary_data.append({
        '項目': '已配對的河川',
        '數量': len(rivers_with_stations_df),
        '百分比': f'{len(rivers_with_stations_df)/len(rivers)*100:.1f}%',
        '備註': '有至少1個測站配對'
    })
    summary_data.append({
        '項目': '未配對的河川',
        '數量': len(rivers_without_stations_df),
        '百分比': f'{len(rivers_without_stations_df)/len(rivers)*100:.1f}%',
        '備註': '有待進一步確認，見「未配對的河川」工作表'
    })

    # 3. 按階層統計
    summary_data.append({'項目': '', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({'項目': '【按河川階層統計】', '數量': '', '百分比': '', '備註': ''})
    for level in sorted(rivers['階層'].unique()):
        rivers_at_level = rivers[rivers['階層'] == level]
        with_stations = rivers_at_level[rivers_at_level['測站數量'] > 0]
        summary_data.append({
            '項目': f'  階層{level}河川',
            '數量': len(rivers_at_level),
            '百分比': '',
            '備註': f'{len(with_stations)}/{len(rivers_at_level)} 有測站 ({len(with_stations)/len(rivers_at_level)*100:.1f}%)'
        })

    # 4. 無法配對原因
    summary_data.append({'項目': '', '數量': '', '百分比': '', '備註': ''})
    summary_data.append({'項目': '【無法配對原因】', '數量': '', '百分比': '', '備註': ''})
    for reason, count in unmatched_stations['無法配對原因'].value_counts().items():
        summary_data.append({
            '項目': f'  {reason}',
            '數量': count,
            '百分比': f'{count/len(unmatched_stations)*100:.1f}%',
            '備註': ''
        })

    summary_df = pd.DataFrame(summary_data)

    # 準備輸出資料
    output_path = Path('data/測站河川配對分析報表.xlsx')
    print(f"\n正在產生報表: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 工作表1: 摘要統計
        summary_df.to_excel(writer, sheet_name='摘要統計', index=False)

        # 工作表2: 能配對的測站
        matched_cols = ['測站名稱', '測站編號', '編號', '測站類型', '水系', '河川',
                       '匹配的河川', '河川代碼', '匹配方式', '管理單位', '高程(m)']
        matched_cols = [c for c in matched_cols if c in matched_stations.columns]
        matched_output = matched_stations[matched_cols].copy()

        # 只按存在的欄位排序
        sort_cols = [c for c in ['水系', '匹配的河川', '測站名稱'] if c in matched_output.columns]
        if sort_cols:
            matched_output = matched_output.sort_values(sort_cols)

        matched_output.to_excel(writer, sheet_name='能配對的測站', index=False)

        # 工作表3: 無法配對的測站
        unmatched_cols = ['測站名稱', '測站編號', '編號', '測站類型', '水系', '河川',
                         '無法配對原因', '管理單位', '高程(m)']
        unmatched_cols = [c for c in unmatched_cols if c in unmatched_stations.columns]
        unmatched_output = unmatched_stations[unmatched_cols].copy()

        # 只按存在的欄位排序
        sort_cols = [c for c in ['無法配對原因', '測站名稱'] if c in unmatched_output.columns]
        if sort_cols:
            unmatched_output = unmatched_output.sort_values(sort_cols)

        unmatched_output.to_excel(writer, sheet_name='無法配對的測站', index=False)

        # 工作表4: 已配對的河川
        river_cols = ['河川名稱', '河川代碼', '階層', '主流水系', '測站數量']
        rivers_with_stations_df[river_cols].to_excel(writer, sheet_name='已配對的河川', index=False)

        # 工作表5: 未配對的河川（有待確認）
        river_cols_no_count = ['河川名稱', '河川代碼', '階層', '上游河川', '主流水系']
        rivers_without_stations_df[river_cols_no_count].to_excel(writer, sheet_name='未配對的河川', index=False)

    # 設定凍結窗格
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
    wb.save(output_path)

    print(f"\n報表已產生: {output_path}")
    print("\n報表內容:")
    print("  - 工作表1: 摘要統計")
    print(f"  - 工作表2: 能配對的測站 ({len(matched_stations)} 個)")
    print(f"  - 工作表3: 無法配對的測站 ({len(unmatched_stations)} 個)")
    print(f"  - 工作表4: 已配對的河川 ({len(rivers_with_stations_df)} 條)")
    print(f"  - 工作表5: 未配對的河川 ({len(rivers_without_stations_df)} 條) [有待確認]")

    # 顯示摘要
    print("\n" + "=" * 80)
    print("報表摘要")
    print("=" * 80)
    print(summary_df.to_string(index=False))

    # 顯示別名匹配範例
    alias_matched = matched_stations[matched_stations['匹配方式'] == '別名匹配']
    if len(alias_matched) > 0:
        print("\n" + "=" * 80)
        print(f"別名匹配範例 (共 {len(alias_matched)} 個)")
        print("=" * 80)
        for idx, row in alias_matched.head(10).iterrows():
            station = row.get('測站名稱', 'N/A')
            original = row.get('河川', 'N/A')
            matched = row.get('匹配的河川', 'N/A')
            try:
                print(f"  測站: {station}")
                print(f"    '{original}' -> '{matched}'")
            except UnicodeEncodeError:
                pass

    # 顯示無法配對的測站摘要
    print("\n" + "=" * 80)
    print(f"無法配對的測站摘要 (共 {len(unmatched_stations)} 個)")
    print("=" * 80)
    for reason, count in unmatched_stations['無法配對原因'].value_counts().items():
        print(f"  {reason}: {count} 個")

    print(f"\n完整名單請查看報表: {output_path}")

def main():
    print("=" * 80)
    print("完整配對報表產生器（改進版）")
    print("=" * 80)

    generate_complete_report_improved()

    print("\n" + "=" * 80)
    print("完成！")
    print("=" * 80)

if __name__ == "__main__":
    main()
